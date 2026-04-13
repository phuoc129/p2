from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from datetime import timedelta
from io import BytesIO
from decimal import Decimal
from openpyxl import load_workbook
from apps.authentication.models import User
from apps.product.models import Category, Product
from apps.warehouse.models import ProductStock, ExportReceipt, ExportReceiptItem
from apps.warehouse.repositories import ProductStockRepository, ExportReceiptRepository
from apps.warehouse.services import ImportReceiptService, ExportReceiptService
from apps.order.models import SalesOrder, SalesOrderItem
from apps.order.services import SalesOrderService


class SalesOrderServiceTestCase(TestCase):
    """Test SalesOrder workflows - Đơn hàng bán"""
    
    def setUp(self):
        """Khởi tạo dữ liệu test"""
        # Tạo users
        self.sale_user = User.objects.create_user(
            username='sale01',
            password='Sale@123',
            email='sale@test.com',
            role='SALE'
        )
        self.admin_user = User.objects.create_user(
            username='admin',
            password='Admin@123',
            email='admin@test.com',
            role='ADMIN',
            is_staff=True,
            is_superuser=True
        )
        self.kho_user = User.objects.create_user(
            username='kho01',
            password='Kho@123',
            role='KHO'
        )
        self.ketoan_user = User.objects.create_user(
            username='ketoan01',
            password='KeToan@123',
            role='KE_TOAN'
        )
        
        # Tạo category & products
        self.category = Category.objects.create(name='Vật liệu xây dựng')
        self.product1 = Product.objects.create(
            name='Xi măng Portland',
            base_price=Decimal('50000'),
            base_unit='Bao',
            category=self.category
        )
        self.product2 = Product.objects.create(
            name='Gạch nung',
            base_price=Decimal('3000'),
            base_unit='Cục',
            category=self.category
        )
        
        self.service = SalesOrderService()
        self.import_service = ImportReceiptService()
    
    def _create_stock(self, product, quantity):
        """Helper: Tạo tồn kho cho sản phẩm"""
        stock, _ = ProductStock.objects.get_or_create(
            product=product,
            defaults={'quantity': quantity}
        )
        if stock.quantity != quantity:
            stock.quantity = quantity
            stock.save()
        return stock
    
    # ═══════════════════════════════════════════════════════════════
    # 1. ĐƠN HÀNG - Tạo đơn
    # ═══════════════════════════════════════════════════════════════
    
    def test_sales_order_create_success(self):
        """Test: SALE tạo đơn hàng thành công"""
        # Tạo tồn kho
        self._create_stock(self.product1, Decimal('200'))
        
        items_data = [
            {
                'product_id': str(self.product1.id),
                'quantity': Decimal('50'),
                'unit_price': Decimal('50000'),
            }
        ]
        
        order, errors = self.service.create_order(
            customer_name='Công ty ABC',
            customer_phone='0901234567',
            note='Giao sáng mai',
            items_data=items_data,
            user=self.sale_user
        )
        
        # Assertions
        self.assertIsNotNone(order)
        self.assertIsNone(errors)
        self.assertEqual(order.status, 'CONFIRMED')
        self.assertEqual(order.created_by, self.sale_user)
        self.assertIn('DH-', order.order_code)
        self.assertEqual(order.items.count(), 1)
        self.assertEqual(order.customer_name, 'Công ty ABC')
        self.assertEqual(order.customer_phone, '0901234567')
    
    def test_sales_order_create_multiple_items(self):
        """Test: Tạo đơn hàng với nhiều sản phẩm"""
        # Tạo tồn kho
        self._create_stock(self.product1, Decimal('300'))
        self._create_stock(self.product2, Decimal('10000'))
        
        items_data = [
            {'product_id': str(self.product1.id), 'quantity': Decimal('50'), 'unit_price': Decimal('50000')},
            {'product_id': str(self.product2.id), 'quantity': Decimal('2000'), 'unit_price': Decimal('3000')},
        ]
        
        order, errors = self.service.create_order(
            customer_name='Công ty XYZ',
            customer_phone='0987654321',
            note='',
            items_data=items_data,
            user=self.sale_user
        )
        
        self.assertIsNotNone(order)
        self.assertEqual(order.items.count(), 2)
        self.assertEqual(order.total_amount, Decimal('50000') * Decimal('50') + Decimal('3000') * Decimal('2000'))
    
    def test_sales_order_create_insufficient_stock_error(self):
        """Test: Tạo đơn với số lượng > tồn kho → lỗi"""
        # Chỉ có 50, nhưng yêu cầu 100
        self._create_stock(self.product1, Decimal('50'))
        
        items_data = [
            {'product_id': str(self.product1.id), 'quantity': Decimal('100'), 'unit_price': Decimal('50000')}
        ]
        
        order, errors = self.service.create_order(
            customer_name='Công ty',
            customer_phone='',
            note='',
            items_data=items_data,
            user=self.sale_user
        )
        
        self.assertIsNone(order)
        self.assertIsNotNone(errors)
        self.assertTrue(len(errors) > 0)
        self.assertIn('chỉ còn 50', errors[0]['message'])
    
    def test_sales_order_create_no_customer_name_error(self):
        """Test: Tạo đơn mà không nhập tên khách → lỗi"""
        self._create_stock(self.product1, Decimal('200'))
        
        items_data = [{'product_id': str(self.product1.id), 'quantity': Decimal('50'), 'unit_price': Decimal('50000')}]
        
        order, errors = self.service.create_order(
            customer_name='',
            customer_phone='',
            note='',
            items_data=items_data,
            user=self.sale_user
        )
        
        self.assertIsNone(order)
        self.assertIsNotNone(errors)
        self.assertIn('tên khách hàng', errors[0]['message'])
    
    def test_sales_order_create_no_items_error(self):
        """Test: Tạo đơn mà không chọn sản phẩm → lỗi"""
        order, errors = self.service.create_order(
            customer_name='Công ty',
            customer_phone='',
            note='',
            items_data=[],
            user=self.sale_user
        )
        
        self.assertIsNone(order)
        self.assertIsNotNone(errors)
        self.assertIn('ít nhất 1', errors[0]['message'])
    
    def test_sales_order_create_invalid_quantity_error(self):
        """Test: Tạo đơn với số lượng không hợp lệ → lỗi"""
        self._create_stock(self.product1, Decimal('200'))
        
        items_data = [{'product_id': str(self.product1.id), 'quantity': 0, 'unit_price': Decimal('50000')}]
        
        order, errors = self.service.create_order(
            customer_name='Công ty',
            customer_phone='',
            note='',
            items_data=items_data,
            user=self.sale_user
        )
        
        self.assertIsNone(order)
        self.assertIsNotNone(errors)
        self.assertIn('lớn hơn 0', errors[0]['message'])
    
    # ═══════════════════════════════════════════════════════════════
    # 2. ĐƠN HÀNG - Cập nhật trạng thái
    # ═══════════════════════════════════════════════════════════════
    
    def test_sales_order_update_to_waiting(self):
        """Test: Chuyển đơn sang WAITING → tạo phiếu xuất tự động"""
        # Setup
        self._create_stock(self.product1, Decimal('200'))
        items_data = [{'product_id': str(self.product1.id), 'quantity': Decimal('50'), 'unit_price': Decimal('50000')}]
        order, _ = self.service.create_order('Công ty', '', '', items_data, self.sale_user)
        
        # Kiểm tra trước khi chuyển
        exports_before = ExportReceipt.objects.filter(note__icontains=order.order_code).count()
        self.assertEqual(exports_before, 0)
        
        # Chuyển sang WAITING
        success, msg = self.service.update_status(order.id, 'WAITING', updated_by=self.admin_user)
        
        # Assertions
        self.assertTrue(success)
        order.refresh_from_db()
        self.assertEqual(order.status, 'WAITING')
        
        # Kiểm tra phiếu xuất tạo tự động
        export_receipts = ExportReceipt.objects.filter(note__icontains=order.order_code)
        self.assertEqual(export_receipts.count(), 1)
        
        export = export_receipts.first()
        self.assertEqual(export.status, 'PENDING')
        self.assertEqual(export.items.count(), 1)
        
        export_item = export.items.first()
        self.assertEqual(export_item.product, self.product1)
        self.assertEqual(export_item.quantity, Decimal('50'))
    
    def test_sales_order_update_status_invalid_transition(self):
        """Test: Chuyển trạng thái không hợp lệ → lỗi"""
        self._create_stock(self.product1, Decimal('200'))
        items_data = [{'product_id': str(self.product1.id), 'quantity': Decimal('50'), 'unit_price': Decimal('50000')}]
        order, _ = self.service.create_order('Công ty', '', '', items_data, self.sale_user)
        
        # Cố chuyển CONFIRMED → DONE (không hợp lệ, phải qua WAITING)
        success, msg = self.service.update_status(order.id, 'DONE')
        
        self.assertFalse(success)
        self.assertIn('Không thể chuyển', msg)
    
    def test_sales_order_cancel(self):
        """Test: Hủy đơn hàng"""
        self._create_stock(self.product1, Decimal('200'))
        items_data = [{'product_id': str(self.product1.id), 'quantity': Decimal('50'), 'unit_price': Decimal('50000')}]
        order, _ = self.service.create_order('Công ty', '', '', items_data, self.sale_user)
        
        # Hủy
        success, msg = self.service.update_status(order.id, 'CANCELLED')
        
        self.assertTrue(success)
        order.refresh_from_db()
        self.assertEqual(order.status, 'CANCELLED')


class ExportReceiptAutoCreateTestCase(TestCase):
    """Test ExportReceipt auto-create & workflows - Phiếu xuất kho tự động"""
    
    def setUp(self):
        """Khởi tạo dữ liệu test"""
        self.sale_user = User.objects.create_user(username='sale01', password='Sale@123', role='SALE')
        self.kho_user = User.objects.create_user(username='kho01', password='Kho@123', role='KHO')
        self.ketoan_user = User.objects.create_user(username='ketoan01', password='KeToan@123', role='KE_TOAN')
        self.admin_user = User.objects.create_user(username='admin', password='Admin@123', role='ADMIN', is_superuser=True, is_staff=True)
        
        self.category = Category.objects.create(name='Vật liệu')
        self.product = Product.objects.create(
            name='Xi măng',
            base_price=Decimal('50000'),
            base_unit='Bao',
            category=self.category
        )
        
        self.sales_service = SalesOrderService()
        self.export_service = ExportReceiptService()
    
    def _create_stock(self, product, quantity):
        """Helper: Tạo tồn kho"""
        stock, _ = ProductStock.objects.get_or_create(
            product=product,
            defaults={'quantity': quantity}
        )
        if stock.quantity != quantity:
            stock.quantity = quantity
            stock.save()
        return stock
    
    def _create_sales_order(self, product, quantity):
        """Helper: Tạo đơn hàng"""
        self._create_stock(product, quantity + Decimal('100'))  # Tồn kho đủ
        items_data = [{'product_id': str(product.id), 'quantity': quantity, 'unit_price': Decimal('50000')}]
        order, _ = self.sales_service.create_order('Công ty', '', '', items_data, self.sale_user)
        return order
    
    def test_export_auto_create_when_order_to_waiting(self):
        """Test: Phiếu xuất tạo tự động khi đơn chuyển sang WAITING"""
        order = self._create_sales_order(self.product, Decimal('50'))
        
        # Kiểm tra chưa có phiếu xuất
        exports = ExportReceipt.objects.filter(note__icontains=order.order_code)
        self.assertEqual(exports.count(), 0)
        
        # Chuyển sang WAITING
        self.sales_service.update_status(order.id, 'WAITING', updated_by=self.admin_user)
        
        # Kiểm tra phiếu xuất được tạo
        exports = ExportReceipt.objects.filter(note__icontains=order.order_code)
        self.assertEqual(exports.count(), 1)
        
        export = exports.first()
        self.assertEqual(export.status, 'PENDING')
        self.assertEqual(export.created_by, self.admin_user)
    
    def test_export_approve_decreases_stock(self):
        """Test: Duyệt phiếu xuất → tồn kho GIẢM + đơn hàng DONE"""
        order = self._create_sales_order(self.product, Decimal('50'))
        
        # Tồn kho trước duyệt = 150 (50 + 100 từ helper)
        stock_before = ProductStockRepository.get_stock(self.product.id)
        self.assertEqual(stock_before.quantity, Decimal('150'))
        
        # Chuyển sang WAITING → tạo phiếu xuất
        self.sales_service.update_status(order.id, 'WAITING', updated_by=self.admin_user)
        export = ExportReceipt.objects.filter(note__icontains=order.order_code).first()
        
        # Duyệt phiếu xuất (dùng service)
        success, msg = self.export_service.approve_receipt(export.id, self.ketoan_user)
        self.assertTrue(success)
        
        # Kiểm tra tồn kho giảm (150 - 50 = 100)
        stock_after = ProductStockRepository.get_stock(self.product.id)
        self.assertEqual(stock_after.quantity, Decimal('100'))
        
        # Kiểm tra đơn hàng → DONE
        order.refresh_from_db()
        self.assertEqual(order.status, 'DONE')
    
    def test_export_reject_returns_order_to_confirmed(self):
        """Test: Từ chối phiếu xuất → đơn hàng quay lại CONFIRMED"""
        order = self._create_sales_order(self.product, Decimal('50'))
        
        # Chuyển sang WAITING
        self.sales_service.update_status(order.id, 'WAITING', updated_by=self.admin_user)
        export = ExportReceipt.objects.get(note__icontains=order.order_code)
        
        # Từ chối (dùng service)
        success, msg = self.export_service.reject_receipt(export.id, self.ketoan_user, 'Lỗi khi kiểm kho')
        self.assertTrue(success)
        
        # Kiểm tra đơn hàng quay lại CONFIRMED
        order.refresh_from_db()
        self.assertEqual(order.status, 'CONFIRMED')
        
        export.refresh_from_db()
        self.assertEqual(export.status, 'REJECTED')
    
    def test_export_archive_order_code_extraction(self):
        """Test: Trích xuất mã đơn hàng từ note của phiếu xuất"""
        order = self._create_sales_order(self.product, Decimal('50'))
        
        # Chuyển sang WAITING
        self.sales_service.update_status(order.id, 'WAITING', updated_by=self.admin_user)
        export = ExportReceipt.objects.get(note__icontains=order.order_code)
        
        # Kiểm tra note chứa mã đơn hàng
        self.assertIn(order.order_code, export.note)
        
        # Kiểm tra extraction
        extracted_code = ExportReceiptRepository._extract_order_code_from_note(export.note)
        self.assertEqual(extracted_code, order.order_code)


class EndToEndWorkflowTestCase(TestCase):
    """Test End-to-End: Toàn bộ luồng từ nhập kho → bán hàng → xuất kho"""
    
    def setUp(self):
        """Khởi tạo dữ liệu test"""
        self.kho_user = User.objects.create_user(username='kho01', password='Kho@123', role='KHO')
        self.ketoan_user = User.objects.create_user(username='ketoan01', password='KeToan@123', role='KE_TOAN')
        self.sale_user = User.objects.create_user(username='sale01', password='Sale@123', role='SALE')
        self.admin_user = User.objects.create_user(username='admin', password='Admin@123', role='ADMIN', is_superuser=True, is_staff=True)
        
        self.category = Category.objects.create(name='Vật liệu')
        self.product = Product.objects.create(
            name='Xi măng Portland',
            base_price=Decimal('50000'),
            base_unit='Bao',
            category=self.category
        )
        
        self.import_service = ImportReceiptService()
        self.sales_service = SalesOrderService()
    
    def test_full_workflow_import_to_sale_to_export(self):
        """Test: Luồng đầy đủ - Nhập → Duyệt → Bán → Xuất → Duyệt"""
        export_service = ExportReceiptService()
        
        # ===== STEP 1: NHẬP KHO =====
        import_items = [{'product_id': str(self.product.id), 'quantity': Decimal('200'), 'unit_price': Decimal('50000')}]
        import_receipt, _ = self.import_service.create_receipt('Nhập lô A', import_items, self.kho_user)
        self.assertEqual(import_receipt.status, 'PENDING')
        
        # Tồn kho = 0 trước duyệt
        stock = ProductStockRepository.get_stock(self.product.id)
        self.assertIsNone(stock)
        
        # Duyệt phiếu nhập
        self.import_service.approve_receipt(import_receipt.id, self.ketoan_user)
        
        # Tồn kho = 200 sau duyệt
        stock = ProductStockRepository.get_stock(self.product.id)
        self.assertEqual(stock.quantity, Decimal('200'))
        
        # ===== STEP 2: SALE TẠO ĐƠN =====
        sales_items = [{'product_id': str(self.product.id), 'quantity': Decimal('50'), 'unit_price': Decimal('50000')}]
        order, _ = self.sales_service.create_order(
            customer_name='Công ty ABC',
            customer_phone='0901234567',
            note='Giao sáng mai',
            items_data=sales_items,
            user=self.sale_user
        )
        self.assertEqual(order.status, 'CONFIRMED')
        
        # Tồn kho vẫn 200 (chưa trừ, vì chỉ trừ khi xuất được duyệt)
        stock = ProductStockRepository.get_stock(self.product.id)
        self.assertEqual(stock.quantity, Decimal('200'))
        
        # ===== STEP 3: CHUYỂN SANG WAITING & XUẤT TỰ ĐỘNG =====
        self.sales_service.update_status(order.id, 'WAITING', updated_by=self.admin_user)
        
        order.refresh_from_db()
        self.assertEqual(order.status, 'WAITING')
        
        # Phiếu xuất được tạo tự động
        export = ExportReceipt.objects.get(note__icontains=order.order_code)
        self.assertEqual(export.status, 'PENDING')
        
        # ===== STEP 4: DUYỆT PHIẾU XUẤT =====
        success, msg = export_service.approve_receipt(export.id, self.ketoan_user)
        self.assertTrue(success)
        
        # Tồn kho = 150 (200 - 50)
        stock = ProductStockRepository.get_stock(self.product.id)
        self.assertEqual(stock.quantity, Decimal('150'))
        
        # Đơn hàng → DONE
        order.refresh_from_db()
        self.assertEqual(order.status, 'DONE')
        
        # Phiếu xuất → APPROVED
        export.refresh_from_db()
        self.assertEqual(export.status, 'APPROVED')


class SalesOrderExportExcelViewTestCase(TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_user(
            username='admin_export_order',
            password='Admin@123',
            role='ADMIN',
            is_staff=True,
            is_superuser=True,
        )
        self.sale_user = User.objects.create_user(
            username='sale_export_order',
            password='Sale@123',
            role='SALE',
        )

        self.category = Category.objects.create(name='Danh muc export')
        self.product_a = Product.objects.create(
            name='San pham A',
            base_price=Decimal('50000'),
            base_unit='Bao',
            category=self.category,
        )
        self.product_b = Product.objects.create(
            name='San pham B',
            base_price=Decimal('30000'),
            base_unit='Thung',
            category=self.category,
        )

        now = timezone.now()
        self.order_confirmed = self._create_order(
            order_code='DH-EXCEL-001',
            customer_name='Cong ty A',
            status='CONFIRMED',
            created_at=now - timedelta(days=3),
        )
        self.order_done = self._create_order(
            order_code='DH-EXCEL-002',
            customer_name='Cong ty B',
            status='DONE',
            created_at=now - timedelta(days=1),
        )

    def _create_order(self, order_code, customer_name, status, created_at):
        order = SalesOrder.objects.create(
            order_code=order_code,
            customer_name=customer_name,
            customer_phone='0900000000',
            created_by=self.sale_user,
            status=status,
        )
        SalesOrderItem.objects.create(
            order=order,
            product=self.product_a,
            quantity=Decimal('2'),
            unit_price=Decimal('50000'),
        )
        SalesOrderItem.objects.create(
            order=order,
            product=self.product_b,
            quantity=Decimal('1.5'),
            unit_price=Decimal('30000'),
        )
        SalesOrder.objects.filter(id=order.id).update(created_at=created_at)
        order.refresh_from_db()
        return order

    def test_export_excel_returns_xlsx_file(self):
        self.client.login(username='admin_export_order', password='Admin@123')
        today = timezone.localdate()
        response = self.client.get(
            reverse('order:sales_export_excel'),
            {
                'from_date': (today - timedelta(days=7)).isoformat(),
                'to_date': today.isoformat(),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', response['Content-Type'])
        self.assertIn('.xlsx', response['Content-Disposition'])

        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active
        self.assertEqual(sheet['A1'].value, 'BAO CAO DON HANG')

        order_codes = [
            sheet.cell(row=row, column=2).value
            for row in range(10, sheet.max_row)
            if sheet.cell(row=row, column=2).value
        ]
        self.assertIn(self.order_confirmed.order_code, order_codes)
        self.assertIn(self.order_done.order_code, order_codes)
        self.assertEqual(sheet['I10'].number_format, '#,##0.##')
        self.assertEqual(sheet['J10'].number_format, '#,##0.##')

    def test_export_excel_applies_status_filter(self):
        self.client.login(username='admin_export_order', password='Admin@123')
        today = timezone.localdate()
        response = self.client.get(
            reverse('order:sales_export_excel'),
            {
                'status': 'DONE',
                'from_date': (today - timedelta(days=7)).isoformat(),
                'to_date': today.isoformat(),
            },
        )

        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active
        order_codes = [
            sheet.cell(row=row, column=2).value
            for row in range(10, sheet.max_row)
            if sheet.cell(row=row, column=2).value
        ]

        self.assertIn(self.order_done.order_code, order_codes)
        self.assertNotIn(self.order_confirmed.order_code, order_codes)

    def test_export_excel_blocks_invalid_date_range(self):
        self.client.login(username='admin_export_order', password='Admin@123')
        today = timezone.localdate()

        response = self.client.get(
            reverse('order:sales_export_excel'),
            {
                'from_date': (today + timedelta(days=1)).isoformat(),
                'to_date': today.isoformat(),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('order:sales_list'), response['Location'])


class SalesOrderExportPdfViewTestCase(TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_user(
            username='admin_pdf_order',
            password='Admin@123',
            role='ADMIN',
            is_staff=True,
            is_superuser=True,
        )
        self.sale_user = User.objects.create_user(
            username='sale_pdf_order',
            password='Sale@123',
            role='SALE',
        )

        self.category = Category.objects.create(name='Danh muc PDF')
        self.product = Product.objects.create(
            name='San pham PDF',
            base_price=Decimal('45000'),
            base_unit='Bao',
            category=self.category,
        )

        now = timezone.now()
        self.order_done = SalesOrder.objects.create(
            order_code='DH-PDF-001',
            customer_name='Cong ty PDF',
            customer_phone='0901231234',
            created_by=self.sale_user,
            status='DONE',
        )
        SalesOrderItem.objects.create(
            order=self.order_done,
            product=self.product,
            quantity=Decimal('3'),
            unit_price=Decimal('45000'),
        )
        SalesOrder.objects.filter(id=self.order_done.id).update(created_at=now - timedelta(days=1))

    def test_export_pdf_returns_pdf_file(self):
        self.client.login(username='admin_pdf_order', password='Admin@123')
        today = timezone.localdate()

        response = self.client.get(
            reverse('order:sales_export_pdf'),
            {
                'from_date': (today - timedelta(days=7)).isoformat(),
                'to_date': today.isoformat(),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('application/pdf', response['Content-Type'])
        self.assertIn('.pdf', response['Content-Disposition'])
        self.assertTrue(response.content.startswith(b'%PDF'))

    def test_export_pdf_blocks_invalid_date_range(self):
        self.client.login(username='admin_pdf_order', password='Admin@123')
        today = timezone.localdate()

        response = self.client.get(
            reverse('order:sales_export_pdf'),
            {
                'from_date': (today + timedelta(days=1)).isoformat(),
                'to_date': today.isoformat(),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('order:sales_list'), response['Location'])

