import json
import os
from datetime import datetime
from decimal import Decimal
from urllib.parse import urlencode

from django.shortcuts import render, redirect
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Sum
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.urls import reverse

from apps.product.models import Product
from .services import SalesOrderService
from .models import SalesOrder, SalesOrderItem


PAGE_SIZE = 5


def _products_json():
    products = Product.objects.select_related('category').all().order_by('name')
    return [
        {
            'id': str(p.id),
            'name': p.name,
            'base_unit': p.base_unit,
            'base_price': float(p.base_price),
            'category': p.category.name if p.category else '',
        }
        for p in products
    ]


def _stocks_json():
    from apps.warehouse.repositories import ProductStockRepository
    stocks = ProductStockRepository.get_all()
    return {str(s.product_id): float(s.available_quantity) for s in stocks}


def _parse_items_from_post(post_data):
    items = []
    i = 0
    while True:
        product_id = post_data.get(f'product_id_{i}')
        if product_id is None:
            break
        if product_id:
            items.append({
                'product_id': product_id,
                'quantity': post_data.get(f'quantity_{i}', 0),
                'unit_price': post_data.get(f'unit_price_{i}', 0),
                'note': post_data.get(f'item_note_{i}', ''),
            })
        i += 1
    return items


def _get_sales_order_stats():
    today = timezone.now().date()
    return {
        'total_orders': SalesOrder.objects.count(),
        'pending_orders': SalesOrder.objects.filter(status='WAITING').count(),
        'total_items': SalesOrderItem.objects.aggregate(total=Sum('quantity'))['total'] or 0,
        'today_transactions': SalesOrder.objects.filter(created_at__date=today).count(),
    }


def _get_user_display_name(user):
    full_name = ''
    if hasattr(user, 'get_full_name'):
        full_name = (user.get_full_name() or '').strip()
    return full_name or getattr(user, 'username', '') or 'Khong ro'


def _parse_sales_report_filters(request):
    today = timezone.localdate()
    first_day = today.replace(day=1)

    status_filter = request.GET.get('status', '').strip()
    search_query = request.GET.get('search', '').strip()
    from_date_str = request.GET.get('from_date', first_day.isoformat())
    to_date_str = request.GET.get('to_date', today.isoformat())

    try:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    except ValueError:
        from_date = first_day
        to_date = today
        messages.error(request, 'Khoảng thời gian không hợp lệ. Hệ thống đã dùng mặc định tháng hiện tại.')

    is_valid_range = True
    if from_date > to_date:
        is_valid_range = False
        messages.error(request, 'Ngày bắt đầu không được lớn hơn ngày kết thúc. Vui lòng chọn lại khoảng thời gian.')

    return status_filter, search_query, from_date, to_date, is_valid_range


def _get_base_orders_for_user(service, user):
    if user.role in ('KE_TOAN', 'ADMIN'):
        return service.get_all()
    if user.role == 'SALE':
        return service.get_by_user(user)
    return service.get_all()


def _apply_sales_order_filters(orders, status_filter, search_query, from_date, to_date):
    if status_filter:
        orders = orders.filter(status=status_filter)

    if search_query:
        orders = orders.filter(
            Q(customer_name__icontains=search_query) |
            Q(order_code__icontains=search_query)
        )

    return orders.filter(
        created_at__date__gte=from_date,
        created_at__date__lte=to_date,
    )


def _format_report_number(value):
    decimal_value = Decimal(str(value))
    formatted = f"{decimal_value:,.10f}".rstrip('0').rstrip('.')
    return formatted or '0'


class SalesOrderListView(LoginRequiredMixin, View):
    """Danh sách đơn hàng"""

    def get(self, request):
        service = SalesOrderService()
        user = request.user
        status_filter, search_query, from_date, to_date, is_valid_range = _parse_sales_report_filters(request)
        page_number = request.GET.get('page', 1)

        orders = _get_base_orders_for_user(service, user)
        if is_valid_range:
            orders = _apply_sales_order_filters(orders, status_filter, search_query, from_date, to_date)
        else:
            orders = orders.none()

        paginator = Paginator(orders, PAGE_SIZE)
        page_obj = paginator.get_page(page_number)

        products_data = _products_json()
        stocks_data = _stocks_json()
        stats = _get_sales_order_stats()

        valid_transitions = SalesOrderService.VALID_TRANSITIONS
        user_role = 'ADMIN' if user.is_superuser else user.role

        return render(request, 'order/sales_order_list.html', {
            'orders': page_obj,
            'page_obj': page_obj,
            'paginator': paginator,
            'products_json': json.dumps(products_data, ensure_ascii=False),
            'stocks_json': json.dumps(stocks_data),
            'user_role': user_role,
            'status_filter': status_filter,
            'search_query': search_query,
            'from_date': from_date.isoformat(),
            'to_date': to_date.isoformat(),
            'stats': stats,
            'valid_transitions_json': json.dumps(valid_transitions),
        })

    def post(self, request):
        user = request.user
        action = request.POST.get('action', '')

        if action == 'update_status':
            if user.role == 'SALE' and not user.is_superuser:
                messages.error(request, 'Bạn không có quyền cập nhật trạng thái đơn hàng.')
                return redirect('order:sales_list')

            order_id = request.POST.get('order_id')
            new_status = request.POST.get('status')

            allowed_statuses = ['CONFIRMED', 'WAITING', 'DONE', 'CANCELLED']
            if new_status not in allowed_statuses:
                messages.error(request, 'Trạng thái không hợp lệ.')
                return redirect('order:sales_list')

            service = SalesOrderService()
            success, msg = service.update_status(order_id, new_status, updated_by=user)
            if success:
                if new_status == 'WAITING':
                    messages.success(request, f'{msg} Phiếu xuất kho đã được tạo tự động và đang chờ duyệt.')
                else:
                    messages.success(request, msg)
            else:
                messages.error(request, msg)
            return redirect('order:sales_list')

        if user.role not in ('SALE', 'ADMIN') and not user.is_superuser:
            messages.error(request, 'Bạn không có quyền tạo đơn hàng.')
            return redirect('order:sales_list')

        service = SalesOrderService()
        customer_name = request.POST.get('customer_name', '')
        customer_phone = request.POST.get('customer_phone', '')
        note = request.POST.get('note', '')
        items_data = _parse_items_from_post(request.POST)

        order, errors = service.create_order(customer_name, customer_phone, note, items_data, user)

        if order:
            messages.success(request, f'Đơn hàng {order.order_code} đã được tạo thành công.')
        else:
            for err in errors:
                messages.error(request, err['message'])

        return redirect('order:sales_list')


class SalesOrderExportExcelView(LoginRequiredMixin, View):
    def get(self, request):
        service = SalesOrderService()
        status_filter, search_query, from_date, to_date, is_valid_range = _parse_sales_report_filters(request)

        if not is_valid_range:
            params = {
                'status': status_filter,
                'search': search_query,
                'from_date': from_date.isoformat(),
                'to_date': to_date.isoformat(),
            }
            params = {k: v for k, v in params.items() if v}
            return redirect(f"{reverse('order:sales_list')}?{urlencode(params)}")

        orders = _get_base_orders_for_user(service, request.user)
        orders = _apply_sales_order_filters(orders, status_filter, search_query, from_date, to_date)

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Bao cao don hang'

        status_labels = dict(SalesOrder.STATUS_CHOICES)
        status_label = status_labels.get(status_filter, status_filter) if status_filter else 'Tat ca trang thai'
        search_label = search_query if search_query else 'Khong co'

        sheet.merge_cells('A1:J1')
        sheet['A1'] = 'BAO CAO DON HANG'
        sheet['A1'].font = Font(size=14, bold=True)
        sheet['A1'].alignment = Alignment(horizontal='center')

        sheet['A2'] = f'Tu ngay: {from_date.strftime("%d/%m/%Y")}'
        sheet['A3'] = f'Den ngay: {to_date.strftime("%d/%m/%Y")}'
        sheet['A4'] = f'Trang thai: {status_label}'
        sheet['A5'] = f'Tu khoa tim kiem: {search_label}'
        sheet['A6'] = f'Xuat luc: {timezone.localtime().strftime("%d/%m/%Y %H:%M:%S")}'
        sheet['A7'] = f'Nguoi xuat: {_get_user_display_name(request.user)}'

        headers = [
            'STT',
            'Ma don',
            'Khach hang',
            'So dien thoai',
            'Nhan vien tao',
            'Trang thai',
            'Ngay tao',
            'So dong SP',
            'Tong so luong',
            'Tong tien',
        ]

        header_row = 9
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(fill_type='solid', fgColor='1F4E78')
            cell.alignment = Alignment(horizontal='center')

        number_format = '#,##0.##'
        total_quantity = Decimal('0')
        total_amount = Decimal('0')
        current_row = header_row + 1

        for index, order in enumerate(orders, start=1):
            order_items = list(order.items.all())
            item_count = len(order_items)
            quantity_sum = sum((item.quantity for item in order_items), Decimal('0'))
            amount_sum = sum((item.subtotal for item in order_items), Decimal('0'))

            total_quantity += quantity_sum
            total_amount += amount_sum

            sheet.cell(row=current_row, column=1, value=index)
            sheet.cell(row=current_row, column=2, value=order.order_code)
            sheet.cell(row=current_row, column=3, value=order.customer_name)
            sheet.cell(row=current_row, column=4, value=order.customer_phone or '')
            sheet.cell(row=current_row, column=5, value=_get_user_display_name(order.created_by))
            sheet.cell(row=current_row, column=6, value=order.get_status_display())
            sheet.cell(row=current_row, column=7, value=timezone.localtime(order.created_at).strftime('%d/%m/%Y %H:%M'))
            sheet.cell(row=current_row, column=8, value=item_count)

            quantity_cell = sheet.cell(row=current_row, column=9, value=float(quantity_sum))
            amount_cell = sheet.cell(row=current_row, column=10, value=float(amount_sum))
            quantity_cell.number_format = number_format
            amount_cell.number_format = number_format

            current_row += 1

        total_row = current_row
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=8)
        total_title = sheet.cell(row=total_row, column=1, value='Tong cong')
        total_title.font = Font(bold=True)
        total_title.alignment = Alignment(horizontal='right')

        total_quantity_cell = sheet.cell(row=total_row, column=9, value=float(total_quantity))
        total_amount_cell = sheet.cell(row=total_row, column=10, value=float(total_amount))
        total_quantity_cell.number_format = number_format
        total_amount_cell.number_format = number_format
        total_quantity_cell.font = Font(bold=True)
        total_amount_cell.font = Font(bold=True)

        column_widths = {
            'A': 7,
            'B': 18,
            'C': 28,
            'D': 18,
            'E': 20,
            'F': 16,
            'G': 20,
            'H': 12,
            'I': 16,
            'J': 18,
        }
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width

        filename = f'bao_cao_don_hang_{from_date.isoformat()}_{to_date.isoformat()}.xlsx'
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        workbook.save(response)
        return response


class SalesOrderExportPdfView(LoginRequiredMixin, View):
    def get(self, request):
        service = SalesOrderService()
        status_filter, search_query, from_date, to_date, is_valid_range = _parse_sales_report_filters(request)

        if not is_valid_range:
            params = {
                'status': status_filter,
                'search': search_query,
                'from_date': from_date.isoformat(),
                'to_date': to_date.isoformat(),
            }
            params = {k: v for k, v in params.items() if v}
            return redirect(f"{reverse('order:sales_list')}?{urlencode(params)}")

        orders = _get_base_orders_for_user(service, request.user)
        orders = _apply_sales_order_filters(orders, status_filter, search_query, from_date, to_date)

        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        status_labels = dict(SalesOrder.STATUS_CHOICES)
        status_label = status_labels.get(status_filter, status_filter) if status_filter else 'Tat ca trang thai'
        search_label = search_query if search_query else 'Khong co'

        filename = f'bao_cao_don_hang_{from_date.isoformat()}_{to_date.isoformat()}.pdf'
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(A4),
            leftMargin=20,
            rightMargin=20,
            topMargin=20,
            bottomMargin=20,
        )

        styles = getSampleStyleSheet()
        regular_font = 'Helvetica'
        bold_font = 'Helvetica-Bold'

        regular_font_path = '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'
        bold_font_path = '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'
        if os.path.exists(regular_font_path) and os.path.exists(bold_font_path):
            pdfmetrics.registerFont(TTFont('DejaVuSans', regular_font_path))
            pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', bold_font_path))
            regular_font = 'DejaVuSans'
            bold_font = 'DejaVuSans-Bold'

        title_style = ParagraphStyle(
            'OrderReportTitle',
            parent=styles['Heading1'],
            alignment=TA_CENTER,
            fontName=bold_font,
            fontSize=16,
            spaceAfter=10,
        )
        body_style = ParagraphStyle(
            'OrderReportBody',
            parent=styles['Normal'],
            fontName=regular_font,
            fontSize=10,
        )

        story = [
            Paragraph('BAO CAO DON HANG', title_style),
            Paragraph(f'Tu ngay: {from_date.strftime("%d/%m/%Y")}', body_style),
            Paragraph(f'Den ngay: {to_date.strftime("%d/%m/%Y")}', body_style),
            Paragraph(f'Trang thai: {status_label}', body_style),
            Paragraph(f'Tu khoa tim kiem: {search_label}', body_style),
            Paragraph(f'Xuat luc: {timezone.localtime().strftime("%d/%m/%Y %H:%M:%S")}', body_style),
            Paragraph(f'Nguoi xuat: {_get_user_display_name(request.user)}', body_style),
            Spacer(1, 10),
        ]

        table_data = [[
            'STT',
            'Ma don',
            'Khach hang',
            'Nhan vien tao',
            'Trang thai',
            'Ngay tao',
            'So dong SP',
            'Tong so luong',
            'Tong tien',
        ]]

        total_quantity = Decimal('0')
        total_amount = Decimal('0')

        for index, order in enumerate(orders, start=1):
            order_items = list(order.items.all())
            item_count = len(order_items)
            quantity_sum = sum((item.quantity for item in order_items), Decimal('0'))
            amount_sum = sum((item.subtotal for item in order_items), Decimal('0'))

            total_quantity += quantity_sum
            total_amount += amount_sum

            table_data.append([
                str(index),
                str(order.order_code),
                str(order.customer_name),
                str(_get_user_display_name(order.created_by)),
                str(order.get_status_display()),
                str(timezone.localtime(order.created_at).strftime('%d/%m/%Y %H:%M')),
                str(item_count),
                _format_report_number(quantity_sum),
                _format_report_number(amount_sum),
            ])

        table_data.append([
            '',
            'TONG CONG',
            '',
            '',
            '',
            '',
            '',
            _format_report_number(total_quantity),
            _format_report_number(total_amount),
        ])

        table = Table(
            table_data,
            repeatRows=1,
            colWidths=[30, 90, 130, 115, 90, 95, 70, 90, 90],
        )
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), bold_font),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
            ('FONTNAME', (0, 1), (-1, -2), regular_font),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (6, 1), (-1, -1), 'RIGHT'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F3F6FA')),
            ('FONTNAME', (0, -1), (-1, -1), bold_font),
        ])
        table.setStyle(table_style)

        story.append(table)
        doc.build(story)
        return response


class SalesOrderDetailView(LoginRequiredMixin, View):
    def get(self, request, pk):
        service = SalesOrderService()
        order = service.get_by_id(pk)
        if not order:
            messages.error(request, 'Không tìm thấy đơn hàng.')
            return redirect('order:sales_list')

        return render(request, 'order/sales_order_detail.html', {
            'order': order,
            'user_role': 'ADMIN' if request.user.is_superuser else request.user.role,
            'valid_transitions': SalesOrderService.VALID_TRANSITIONS.get(order.status, []),
        })

class SalesReportView(LoginRequiredMixin, View):
 
    def get(self, request):
        from django.db.models import (
            F, Sum, Count, ExpressionWrapper, DecimalField
        )
        from django.db.models.functions import TruncMonth, TruncDate
        from apps.order.models import SalesOrderItem
        from datetime import timedelta
        import json
 
        user = request.user
        service = SalesOrderService()
 
        # Phân quyền
        user_role = 'ADMIN' if user.is_superuser else user.role
        if user_role == 'SALE':
            base_qs = service.get_by_user(user)
        else:
            base_qs = service.get_all()
 
        # Bộ lọc ngày (mặc định 12 tháng gần nhất)
        today = timezone.localdate()
        default_from = (today.replace(day=1)).replace(
            year=today.year - 1 if today.month == 1 else today.year,
            month=12 if today.month == 1 else today.month,
        )
        # Đơn giản hơn: lùi 365 ngày
        default_from = today - timedelta(days=364)
        default_from = default_from.replace(day=1)
 
        from_date_str = request.GET.get('from_date', default_from.isoformat())
        to_date_str   = request.GET.get('to_date',   today.isoformat())
 
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
            to_date   = datetime.strptime(to_date_str,   '%Y-%m-%d').date()
        except ValueError:
            from_date = default_from
            to_date   = today
 
        if from_date > to_date:
            from_date, to_date = default_from, today
 
        filtered_qs = base_qs.filter(
            created_at__date__gte=from_date,
            created_at__date__lte=to_date,
        )
 
        # ── Doanh thu & số đơn theo tháng ────────────────────
        monthly_counts = (
            filtered_qs
            .annotate(month=TruncMonth('created_at'))
            .values('month')
            .annotate(order_count=Count('id'))
            .order_by('month')
        )
        count_map = {
            r['month'].strftime('%Y-%m'): r['order_count']
            for r in monthly_counts
        }
 
        revenue_expr = ExpressionWrapper(
            F('quantity') * F('unit_price'),
            output_field=DecimalField(max_digits=20, decimal_places=4)
        )
        monthly_revenue = (
            SalesOrderItem.objects
            .filter(order__in=filtered_qs)
            .annotate(month=TruncMonth('order__created_at'))
            .values('month')
            .annotate(revenue=Sum(revenue_expr))
            .order_by('month')
        )
        revenue_map = {
            r['month'].strftime('%Y-%m'): float(r['revenue'] or 0)
            for r in monthly_revenue
        }
 
        # Tạo dãy tháng liên tiếp
        chart_labels  = []
        chart_revenue = []
        chart_orders  = []
        cur = from_date.replace(day=1)
        end_month = to_date.replace(day=1)
        while cur <= end_month:
            key = cur.strftime('%Y-%m')
            chart_labels.append(cur.strftime('%m/%Y'))
            chart_revenue.append(revenue_map.get(key, 0))
            chart_orders.append(count_map.get(key, 0))
            if cur.month == 12:
                cur = cur.replace(year=cur.year + 1, month=1)
            else:
                cur = cur.replace(month=cur.month + 1)
 
        # ── Thống kê trạng thái ───────────────────────────────
        status_stats = (
            filtered_qs.values('status').annotate(cnt=Count('id'))
        )
        status_map = {s['status']: s['cnt'] for s in status_stats}
 
        total_orders  = filtered_qs.count()
        total_revenue = sum(chart_revenue)
        done_orders   = status_map.get('DONE', 0)
        completion_rate = round(done_orders / total_orders * 100, 1) if total_orders else 0
 
        # ── Top sản phẩm bán chạy (chỉ đơn DONE) ─────────────
        top_products = (
            SalesOrderItem.objects
            .filter(order__in=filtered_qs, order__status='DONE')
            .values('product__name', 'product__base_unit')
            .annotate(
                total_qty=Sum('quantity'),
                total_rev=Sum(revenue_expr),
            )
            .order_by('-total_rev')[:10]
        )
 
        # ── Số đơn theo ngày (30 ngày gần nhất) ──────────────
        thirty_ago = to_date - timedelta(days=29)
        daily_raw = (
            filtered_qs
            .filter(created_at__date__gte=thirty_ago)
            .annotate(day=TruncDate('created_at'))
            .values('day')
            .annotate(cnt=Count('id'))
            .order_by('day')
        )
        daily_map = {r['day'].strftime('%Y-%m-%d'): r['cnt'] for r in daily_raw}
        daily_labels = []
        daily_counts = []
        d = thirty_ago
        while d <= to_date:
            daily_labels.append(d.strftime('%d/%m'))
            daily_counts.append(daily_map.get(d.strftime('%Y-%m-%d'), 0))
            d += timedelta(days=1)
 
        context = {
            'from_date': from_date.isoformat(),
            'to_date':   to_date.isoformat(),
            # Chart JSON
            'chart_labels_json':  json.dumps(chart_labels),
            'chart_revenue_json': json.dumps(chart_revenue),
            'chart_orders_json':  json.dumps(chart_orders),
            'daily_labels_json':  json.dumps(daily_labels),
            'daily_counts_json':  json.dumps(daily_counts),
            # Donut
            'status_confirmed': status_map.get('CONFIRMED', 0),
            'status_waiting':   status_map.get('WAITING', 0),
            'status_done':      status_map.get('DONE', 0),
            'status_cancelled': status_map.get('CANCELLED', 0),
            # KPIs
            'total_orders':     total_orders,
            'total_revenue':    total_revenue,
            'done_orders':      done_orders,
            'cancel_orders':    status_map.get('CANCELLED', 0),
            'pending_orders':   status_map.get('WAITING', 0) + status_map.get('CONFIRMED', 0),
            'completion_rate':  completion_rate,
            # Table
            'top_products':     top_products,
            'user_role':        user_role,
        }
        return render(request, 'order/sales_report.html', context)