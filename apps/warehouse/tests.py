from django.test import TestCase
from django.utils import timezone
from django.urls import reverse
import uuid
from datetime import timedelta
from decimal import Decimal
from io import BytesIO
from openpyxl import load_workbook
from apps.authentication.models import User
from apps.product.models import Category, Product
from apps.warehouse.models import ImportReceipt, ImportReceiptItem, ProductStock, ExportReceipt, ExportReceiptItem
from apps.warehouse.services import ImportReceiptService, StockService, StockReportService, ExportReceiptService
from apps.warehouse.repositories import ImportReceiptRepository, ProductStockRepository, ExportReceiptRepository


class ImportReceiptServiceTestCase(TestCase):
    """Test ImportReceipt workflows - Phiếu nhập kho"""
    
    def setUp(self):
        """Khởi tạo dữ liệu test"""
        # Tạo users
        self.kho_user = User.objects.create_user(
            username='kho01',
            password='Kho@123',
            email='kho@test.com',
            role='KHO'
        )
        self.ketoan_user = User.objects.create_user(
            username='ketoan01',
            password='KeToan@123',
            email='ketoan@test.com',
            role='KE_TOAN'
        )
        
        # Tạo category
        self.category = Category.objects.create(name='Vật liệu xây dựng')
        
        # Tạo 2 sản phẩm
        self.product1 = Product.objects.create(
            name='Xi măng Portland',
            base_price=Decimal('50000'),
            base_unit='Bao',
            category=self.category
        )
        self.product2 = Product.objects.create(
            name='Gạch nung',
            base_price=Decimal('3000'),
            base_unit='Cục',
            category=self.category
        )
        
        self.service = ImportReceiptService()
    
    # ═══════════════════════════════════════════════════════════════
    # 1. NHẬP KHO - Tạo phiếu
    # ═══════════════════════════════════════════════════════════════
    
    def test_import_receipt_create_success(self):
        """Test: KHO tạo phiếu nhập thành công"""
        items_data = [
            {
                'product_id': str(self.product1.id),
                'quantity': Decimal('100'),
                'unit_price': Decimal('50000'),
                'note': 'Nhập từ nhà máy A'
            }
        ]
        
        receipt, error = self.service.create_receipt(
            note='Nhập hàng tháng 4',
            items_data=items_data,
            user=self.kho_user
        )
        
        # Assertions
        self.assertIsNotNone(receipt)
        self.assertIsNone(error)
        self.assertEqual(receipt.status, 'PENDING')
        self.assertEqual(receipt.created_by, self.kho_user)
        self.assertIn('PN-', receipt.receipt_code)
        self.assertEqual(receipt.items.count(), 1)
        
        # Kiểm tra item chi tiết
        item = receipt.items.first()
        self.assertEqual(item.product, self.product1)
        self.assertEqual(item.quantity, Decimal('100'))
        self.assertEqual(item.unit_price, Decimal('50000'))
    
    def test_import_receipt_create_multiple_items(self):
        """Test: KHO tạo phiếu nhập với nhiều sản phẩm"""
        items_data = [
            {
                'product_id': str(self.product1.id),
                'quantity': Decimal('100'),
                'unit_price': Decimal('50000'),
            },
            {
                'product_id': str(self.product2.id),
                'quantity': Decimal('5000'),
                'unit_price': Decimal('3000'),
            }
        ]
        
        receipt, error = self.service.create_receipt(
            note='Nhập hàng lô A',
            items_data=items_data,
            user=self.kho_user
        )
        
        self.assertIsNotNone(receipt)
        self.assertEqual(receipt.items.count(), 2)
        self.assertEqual(receipt.total_items, 2)
    
    def test_import_receipt_create_no_products_error(self):
        """Test: Tạo phiếu nhập mà không chọn sản phẩm → lỗi"""
        items_data = []
        
        receipt, error = self.service.create_receipt(
            note='Phiếu rỗng',
            items_data=items_data,
            user=self.kho_user
        )
        
        self.assertIsNone(receipt)
        self.assertIsNotNone(error)
        self.assertIn('ít nhất 1', error)
    
    def test_import_receipt_create_invalid_quantity_error(self):
        """Test: Tạo phiếu với số lượng không hợp lệ → lỗi"""
        items_data = [
            {
                'product_id': str(self.product1.id),
                'quantity': 'invalid',  # Invalid
                'unit_price': Decimal('50000'),
            }
        ]
        
        receipt, error = self.service.create_receipt(
            note='Phiếu lỗi',
            items_data=items_data,
            user=self.kho_user
        )
        
        self.assertIsNone(receipt)
        self.assertIsNotNone(error)
    
    def test_import_receipt_create_zero_quantity_error(self):
        """Test: Tạo phiếu với số lượng = 0 → lỗi"""
        items_data = [
            {
                'product_id': str(self.product1.id),
                'quantity': 0,
                'unit_price': Decimal('50000'),
            }
        ]
        
        receipt, error = self.service.create_receipt(
            note='Phiếu 0',
            items_data=items_data,
            user=self.kho_user
        )
        
        self.assertIsNone(receipt)
        self.assertIsNotNone(error)
        self.assertIn('lớn hơn 0', error)
    
    # ═══════════════════════════════════════════════════════════════
    # 2. NHẬP KHO - Duyệt phiếu
    # ═══════════════════════════════════════════════════════════════
    
    def test_import_receipt_approve_success(self):
        """Test: KE_TOAN duyệt phiếu → APPROVED + tồn kho TĂNG"""
        # Tạo phiếu
        items_data = [
            {
                'product_id': str(self.product1.id),
                'quantity': Decimal('100'),
                'unit_price': Decimal('50000'),
            }
        ]
        receipt, _ = self.service.create_receipt('Test', items_data, self.kho_user)
        
        # Trước duyệt: tồn kho = 0
        stock_before = ProductStockRepository.get_stock(self.product1.id)
        self.assertIsNone(stock_before)  # Chưa tồn tại
        
        # Duyệt phiếu
        success, msg = self.service.approve_receipt(receipt.id, self.ketoan_user)
        
        # Assertions
        self.assertTrue(success)
        self.assertIn('đã được duyệt', msg)
        
        # Kiểm tra trạng thái phiếu
        receipt.refresh_from_db()
        self.assertEqual(receipt.status, 'APPROVED')
        self.assertEqual(receipt.reviewed_by, self.ketoan_user)
        self.assertIsNotNone(receipt.reviewed_at)
        
        # Kiểm tra tồn kho tăng
        stock_after = ProductStockRepository.get_stock(self.product1.id)
        self.assertIsNotNone(stock_after)
        self.assertEqual(stock_after.quantity, Decimal('100'))
    
    def test_import_receipt_approve_multiple_products(self):
        """Test: Duyệt phiếu nhiều sản phẩm → tồn kho tất cả TĂNG"""
        items_data = [
            {'product_id': str(self.product1.id), 'quantity': Decimal('100'), 'unit_price': Decimal('50000')},
            {'product_id': str(self.product2.id), 'quantity': Decimal('5000'), 'unit_price': Decimal('3000')},
        ]
        receipt, _ = self.service.create_receipt('Test', items_data, self.kho_user)
        
        # Duyệt
        success, _ = self.service.approve_receipt(receipt.id, self.ketoan_user)
        self.assertTrue(success)
        
        # Kiểm tra tồn kho cả 2 sản phẩm
        stock1 = ProductStockRepository.get_stock(self.product1.id)
        stock2 = ProductStockRepository.get_stock(self.product2.id)
        self.assertEqual(stock1.quantity, Decimal('100'))
        self.assertEqual(stock2.quantity, Decimal('5000'))
    
    def test_import_receipt_approve_non_pending_error(self):
        """Test: Duyệt phiếu không ở trạng thái PENDING → lỗi"""
        items_data = [{'product_id': str(self.product1.id), 'quantity': Decimal('100'), 'unit_price': Decimal('50000')}]
        receipt, _ = self.service.create_receipt('Test', items_data, self.kho_user)
        
        # Duyệt lần 1
        self.service.approve_receipt(receipt.id, self.ketoan_user)
        
        # Cố duyệt lần 2 → lỗi
        success, msg = self.service.approve_receipt(receipt.id, self.ketoan_user)
        self.assertFalse(success)
        self.assertIn('chờ duyệt', msg)
    
    # ═══════════════════════════════════════════════════════════════
    # 3. NHẬP KHO - Từ chối phiếu
    # ═══════════════════════════════════════════════════════════════
    
    def test_import_receipt_reject_success(self):
        """Test: KE_TOAN từ chối phiếu → REJECTED + ghi lý do"""
        items_data = [{'product_id': str(self.product1.id), 'quantity': Decimal('100'), 'unit_price': Decimal('50000')}]
        receipt, _ = self.service.create_receipt('Test', items_data, self.kho_user)
        
        # Từ chối
        rejection_note = 'Số lượng không khớp với hóa đơn'
        success, msg = self.service.reject_receipt(receipt.id, self.ketoan_user, rejection_note)
        
        # Assertions
        self.assertTrue(success)
        self.assertIn('bị từ chối', msg)
        
        receipt.refresh_from_db()
        self.assertEqual(receipt.status, 'REJECTED')
        self.assertEqual(receipt.rejection_note, rejection_note)
        self.assertEqual(receipt.reviewed_by, self.ketoan_user)
        
        # Kiểm tra tồn kho KHÔNG thay đổi
        stock = ProductStockRepository.get_stock(self.product1.id)
        self.assertIsNone(stock)  # Vẫn không tạo stock
    
    def test_import_receipt_reject_empty_note_error(self):
        """Test: Từ chối mà không ghi lý do → lỗi"""
        items_data = [{'product_id': str(self.product1.id), 'quantity': Decimal('100'), 'unit_price': Decimal('50000')}]
        receipt, _ = self.service.create_receipt('Test', items_data, self.kho_user)
        
        success, msg = self.service.reject_receipt(receipt.id, self.ketoan_user, '')
        
        self.assertFalse(success)
        self.assertIn('lý do', msg)
    
    # ═══════════════════════════════════════════════════════════════
    # 4. NHẬP KHO - Sửa và gửi lại
    # ═══════════════════════════════════════════════════════════════
    
    def test_import_receipt_resubmit_success(self):
        """Test: KHO sửa phiếu bị từ chối và gửi lại"""
        items_data = [{'product_id': str(self.product1.id), 'quantity': Decimal('100'), 'unit_price': Decimal('50000')}]
        receipt, _ = self.service.create_receipt('Test', items_data, self.kho_user)
        
        # Từ chối
        self.service.reject_receipt(receipt.id, self.ketoan_user, 'Lỗi số lượng')
        
        # Sửa và gửi lại
        new_items = [{'product_id': str(self.product1.id), 'quantity': Decimal('120'), 'unit_price': Decimal('50000')}]
        receipt_resubmitted, error = self.service.resubmit_receipt(
            receipt.id,
            note='Sửa lại phiếu',
            items_data=new_items,
            user=self.kho_user
        )
        
        # Assertions
        self.assertIsNotNone(receipt_resubmitted)
        self.assertIsNone(error)
        self.assertEqual(receipt_resubmitted.status, 'PENDING')
        self.assertEqual(receipt_resubmitted.rejection_note, '')
        self.assertEqual(receipt_resubmitted.reviewed_by, None)
        
        # Kiểm tra items được cập nhật
        item = receipt_resubmitted.items.first()
        self.assertEqual(item.quantity, Decimal('120'))
    
    def test_import_receipt_resubmit_permission_error(self):
        """Test: Người khác không thể sửa phiếu"""
        items_data = [{'product_id': str(self.product1.id), 'quantity': Decimal('100'), 'unit_price': Decimal('50000')}]
        receipt, _ = self.service.create_receipt('Test', items_data, self.kho_user)
        
        # Từ chối
        self.service.reject_receipt(receipt.id, self.ketoan_user, 'Lỗi')
        
        # User khác cố sửa
        other_user = User.objects.create_user(username='kho02', password='Kho@123', role='KHO')
        new_items = [{'product_id': str(self.product1.id), 'quantity': Decimal('120'), 'unit_price': Decimal('50000')}]
        
        receipt_result, error = self.service.resubmit_receipt(receipt.id, 'Test', new_items, other_user)
        
        self.assertIsNone(receipt_result)
        self.assertIsNotNone(error)
        self.assertIn('không có quyền', error)


class ProductStockServiceTestCase(TestCase):
    """Test ProductStock - Tồn kho"""
    
    def setUp(self):
        """Khởi tạo dữ liệu test"""
        # Tạo users
        self.kho_user = User.objects.create_user(username='kho01', password='Kho@123', role='KHO')
        self.ketoan_user = User.objects.create_user(username='ketoan01', password='KeToan@123', role='KE_TOAN')
        
        # Tạo category & products
        self.category = Category.objects.create(name='Vật liệu')
        self.product = Product.objects.create(
            name='Xi măng',
            base_price=Decimal('50000'),
            base_unit='Bao',
            category=self.category
        )
        
        self.service = StockService()
    
    def test_stock_increase_after_import_approve(self):
        """Test: Tồn kho TĂNG sau khi phiếu nhập được duyệt"""
        # Khởi tạo tồn kho = 0
        stock_init = ProductStockRepository.get_stock(self.product.id)
        self.assertIsNone(stock_init)
        
        # Tạo và duyệt phiếu nhập
        import_service = ImportReceiptService()
        items = [{'product_id': str(self.product.id), 'quantity': Decimal('100'), 'unit_price': Decimal('50000')}]
        receipt, _ = import_service.create_receipt('Test', items, self.kho_user)
        import_service.approve_receipt(receipt.id, self.ketoan_user)
        
        # Kiểm tra tồn kho
        stock = self.service.get_stock_info(self.product.id)
        self.assertIsNotNone(stock)
        self.assertEqual(stock.quantity, Decimal('100'))
    
    def test_stock_multiple_imports_cumulative(self):
        """Test: Nhiều lần nhập kho → tồn kho cộng dồn"""
        import_service = ImportReceiptService()
        
        # Lần nhập 1
        items1 = [{'product_id': str(self.product.id), 'quantity': Decimal('100'), 'unit_price': Decimal('50000')}]
        receipt1, _ = import_service.create_receipt('Lần 1', items1, self.kho_user)
        import_service.approve_receipt(receipt1.id, self.ketoan_user)
        
        stock1 = self.service.get_stock_info(self.product.id)
        self.assertEqual(stock1.quantity, Decimal('100'))
        
        # Lần nhập 2
        items2 = [{'product_id': str(self.product.id), 'quantity': Decimal('50'), 'unit_price': Decimal('50000')}]
        receipt2, _ = import_service.create_receipt('Lần 2', items2, self.kho_user)
        import_service.approve_receipt(receipt2.id, self.ketoan_user)
        
        # Kiểm tra cộng dồn
        stock2 = self.service.get_stock_info(self.product.id)
        self.assertEqual(stock2.quantity, Decimal('150'))
    
    def test_get_all_stocks(self):
        """Test: Lấy tồn kho tất cả sản phẩm"""
        import_service = ImportReceiptService()
        
        # Tạo sản phẩm thứ 2
        product2 = Product.objects.create(
            name='Gạch nung',
            base_price=Decimal('3000'),
            base_unit='Cục',
            category=self.category
        )
        
        # Nhập kho cho cả 2
        items1 = [{'product_id': str(self.product.id), 'quantity': Decimal('100'), 'unit_price': Decimal('50000')}]
        receipt1, _ = import_service.create_receipt('Test', items1, self.kho_user)
        import_service.approve_receipt(receipt1.id, self.ketoan_user)
        
        items2 = [{'product_id': str(product2.id), 'quantity': Decimal('5000'), 'unit_price': Decimal('3000')}]
        receipt2, _ = import_service.create_receipt('Test', items2, self.kho_user)
        import_service.approve_receipt(receipt2.id, self.ketoan_user)
        
        # Lấy tất cả tồn kho
        all_stocks = self.service.get_all_stocks()
        self.assertEqual(all_stocks.count(), 2)


class StockReportServiceTestCase(TestCase):
    """Test báo cáo tồn kho theo thời gian (US-23)."""

    def setUp(self):
        self.user = User.objects.create_user(username='admin01', password='Admin@123', role='ADMIN')

        self.cat_a = Category.objects.create(name='Nhóm A')
        self.cat_b = Category.objects.create(name='Nhóm B')

        self.product_a = Product.objects.create(
            name='Sản phẩm A',
            base_price=Decimal('10000'),
            base_unit='Bao',
            category=self.cat_a,
        )
        self.product_b = Product.objects.create(
            name='Sản phẩm B',
            base_price=Decimal('20000'),
            base_unit='Bao',
            category=self.cat_b,
        )

        self.service = StockReportService()

    def _create_import(self, product, quantity, reviewed_at):
        receipt = ImportReceipt.objects.create(
            receipt_code=f'PN-TEST-{uuid.uuid4().hex[:10]}',
            created_by=self.user,
            reviewed_by=self.user,
            status='APPROVED',
            reviewed_at=reviewed_at,
        )
        ImportReceiptItem.objects.create(
            receipt=receipt,
            product=product,
            quantity=Decimal(str(quantity)),
            unit_price=Decimal('0'),
        )

    def _create_export(self, product, quantity, reviewed_at):
        receipt = ExportReceipt.objects.create(
            receipt_code=f'EX-TEST-{uuid.uuid4().hex[:10]}',
            created_by=self.user,
            reviewed_by=self.user,
            status='APPROVED',
            reviewed_at=reviewed_at,
        )
        ExportReceiptItem.objects.create(
            receipt=receipt,
            product=product,
            quantity=Decimal(str(quantity)),
            unit_price=Decimal('0'),
        )

    def test_build_report_by_time_range(self):
        today = timezone.localdate()
        from_date = today - timedelta(days=3)
        to_date = today

        before_period = timezone.now() - timedelta(days=4)
        in_period = timezone.now() - timedelta(days=1)

        # Lịch sử sản phẩm A:
        # Trước kỳ: +100, -30 => tồn đầu = 70
        # Trong kỳ: +20, -10 => tồn cuối = 80
        self._create_import(self.product_a, 100, before_period)
        self._create_export(self.product_a, 30, before_period)
        self._create_import(self.product_a, 20, in_period)
        self._create_export(self.product_a, 10, in_period)

        rows, totals = self.service.build_report(from_date, to_date)

        row_a = next(row for row in rows if row['product'].id == self.product_a.id)
        self.assertEqual(row_a['opening'], Decimal('70'))
        self.assertEqual(row_a['import_qty'], Decimal('20'))
        self.assertEqual(row_a['export_qty'], Decimal('10'))
        self.assertEqual(row_a['closing'], Decimal('80'))

        self.assertEqual(totals['opening'], sum((r['opening'] for r in rows), Decimal('0')))
        self.assertEqual(totals['closing'], sum((r['closing'] for r in rows), Decimal('0')))

    def test_build_report_with_category_filter(self):
        today = timezone.localdate()
        from_date = today - timedelta(days=7)
        to_date = today
        in_period = timezone.now() - timedelta(days=1)

        self._create_import(self.product_a, 50, in_period)
        self._create_import(self.product_b, 70, in_period)

        rows, _ = self.service.build_report(from_date, to_date, category_id=str(self.cat_a.id))

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['product'].id, self.product_a.id)


class StockReportExportExcelViewTestCase(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='admin_export', password='Admin@123', role='ADMIN')
        self.cat_a = Category.objects.create(name='Cat A')
        self.cat_b = Category.objects.create(name='Cat B')
        self.product_a = Product.objects.create(
            name='Vat tu A',
            base_price=Decimal('10000'),
            base_unit='Bao',
            category=self.cat_a,
        )
        self.product_b = Product.objects.create(
            name='Vat tu B',
            base_price=Decimal('20000'),
            base_unit='Bao',
            category=self.cat_b,
        )

    def _create_import(self, product, quantity, reviewed_at):
        receipt = ImportReceipt.objects.create(
            receipt_code=f'PN-EXCEL-{uuid.uuid4().hex[:8]}',
            created_by=self.user,
            reviewed_by=self.user,
            status='APPROVED',
            reviewed_at=reviewed_at,
        )
        ImportReceiptItem.objects.create(
            receipt=receipt,
            product=product,
            quantity=Decimal(str(quantity)),
            unit_price=Decimal('0'),
        )

    def test_export_excel_returns_xlsx_file(self):
        self.client.login(username='admin_export', password='Admin@123')
        reviewed_at = timezone.now() - timedelta(days=1)
        self._create_import(self.product_a, 40, reviewed_at)

        today = timezone.localdate()
        from_date = (today - timedelta(days=7)).isoformat()
        to_date = today.isoformat()

        response = self.client.get(
            reverse('warehouse:stock_report_export_excel'),
            {'from_date': from_date, 'to_date': to_date}
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', response['Content-Type'])
        self.assertIn('.xlsx', response['Content-Disposition'])

        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active

        self.assertEqual(sheet['A1'].value, 'BAO CAO TON KHO THEO THOI GIAN')
        self.assertEqual(sheet['A6'].value, 'Nguoi xuat: admin_export')
        self.assertEqual(sheet['B8'].value, 'Vat tu A')
        self.assertEqual(sheet['E8'].value, 0)
        self.assertEqual(sheet['F8'].value, 40)
        self.assertEqual(sheet['H8'].value, 40)
        self.assertEqual(sheet['F8'].number_format, '#,##0.##')

    def test_export_excel_applies_category_filter(self):
        self.client.login(username='admin_export', password='Admin@123')
        reviewed_at = timezone.now() - timedelta(days=1)
        self._create_import(self.product_a, 10, reviewed_at)
        self._create_import(self.product_b, 20, reviewed_at)

        today = timezone.localdate()
        from_date = (today - timedelta(days=7)).isoformat()
        to_date = today.isoformat()

        response = self.client.get(
            reverse('warehouse:stock_report_export_excel'),
            {
                'from_date': from_date,
                'to_date': to_date,
                'category': str(self.cat_a.id),
            }
        )

        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active

        # Chỉ còn 1 dòng dữ liệu của danh mục Cat A.
        self.assertEqual(sheet['B8'].value, 'Vat tu A')
        self.assertIsNone(sheet['B9'].value)

    def test_stock_report_view_blocks_invalid_date_range(self):
        self.client.login(username='admin_export', password='Admin@123')
        reviewed_at = timezone.now() - timedelta(days=1)
        self._create_import(self.product_a, 15, reviewed_at)

        today = timezone.localdate()
        response = self.client.get(
            reverse('warehouse:stock_report'),
            {
                'from_date': (today + timedelta(days=1)).isoformat(),
                'to_date': today.isoformat(),
            }
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['rows'], [])
        self.assertEqual(response.context['totals']['opening'], 0)
        self.assertEqual(response.context['totals']['import_qty'], 0)
        self.assertEqual(response.context['totals']['export_qty'], 0)
        self.assertEqual(response.context['totals']['closing'], 0)

    def test_export_excel_blocks_invalid_date_range(self):
        self.client.login(username='admin_export', password='Admin@123')
        today = timezone.localdate()

        response = self.client.get(
            reverse('warehouse:stock_report_export_excel'),
            {
                'from_date': (today + timedelta(days=1)).isoformat(),
                'to_date': today.isoformat(),
            }
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('warehouse:stock_report'), response['Location'])


class StockReportExportPdfViewTestCase(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='admin_pdf', password='Admin@123', role='ADMIN')
        self.category = Category.objects.create(name='Cat PDF')
        self.product = Product.objects.create(
            name='Vat tu PDF',
            base_price=Decimal('15000'),
            base_unit='Bao',
            category=self.category,
        )

    def _create_import(self, quantity, reviewed_at):
        receipt = ImportReceipt.objects.create(
            receipt_code=f'PN-PDF-{uuid.uuid4().hex[:8]}',
            created_by=self.user,
            reviewed_by=self.user,
            status='APPROVED',
            reviewed_at=reviewed_at,
        )
        ImportReceiptItem.objects.create(
            receipt=receipt,
            product=self.product,
            quantity=Decimal(str(quantity)),
            unit_price=Decimal('0'),
        )

    def test_export_pdf_returns_pdf_file(self):
        self.client.login(username='admin_pdf', password='Admin@123')
        self._create_import(25, timezone.now() - timedelta(days=1))

        today = timezone.localdate()
        response = self.client.get(
            reverse('warehouse:stock_report_export_pdf'),
            {
                'from_date': (today - timedelta(days=7)).isoformat(),
                'to_date': today.isoformat(),
            }
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('application/pdf', response['Content-Type'])
        self.assertIn('.pdf', response['Content-Disposition'])
        self.assertTrue(response.content.startswith(b'%PDF'))

    def test_export_pdf_blocks_invalid_date_range(self):
        self.client.login(username='admin_pdf', password='Admin@123')
        today = timezone.localdate()

        response = self.client.get(
            reverse('warehouse:stock_report_export_pdf'),
            {
                'from_date': (today + timedelta(days=1)).isoformat(),
                'to_date': today.isoformat(),
            }
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('warehouse:stock_report'), response['Location'])
