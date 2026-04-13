import json
import os
import uuid
from datetime import datetime
from decimal import Decimal
from urllib.parse import urlencode

from django.shortcuts import render, redirect
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.utils import timezone
from django.db.models import Sum, Q
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.urls import reverse

from apps.product.models import Product
from .services import ImportReceiptService, StockService, StockReportService, ExportReceiptService
from .models import ImportReceipt, ImportReceiptItem, ExportReceipt, ExportReceiptItem


def _products_json():
    products = Product.objects.select_related('category').all().order_by('name')
    return [
        {
            'id': str(p.id),
            'name': p.name,
            'base_unit': p.base_unit,
            'base_price': float(p.base_price),
            'category': p.category.name if p.category else '',
        }
        for p in products
    ]


def _parse_items_from_post(post_data):
    items = []
    i = 0
    while True:
        product_id = post_data.get(f'product_id_{i}')
        if product_id is None:
            break
        if product_id:
            items.append({
                'product_id': product_id,
                'quantity': post_data.get(f'quantity_{i}', 0),
                'unit_price': post_data.get(f'unit_price_{i}', 0),
                'note': post_data.get(f'item_note_{i}', ''),
            })
        i += 1
    return items


def _get_import_receipt_stats():
    today = timezone.now().date()
    return {
        'total_receipts': ImportReceipt.objects.count(),
        'pending_receipts': ImportReceipt.objects.filter(status='PENDING').count(),
        'total_items': ImportReceiptItem.objects.aggregate(total=Sum('quantity'))['total'] or 0,
        'today_transactions': ImportReceipt.objects.filter(created_at__date=today).count(),
    }


def _get_export_receipt_stats():
    today = timezone.now().date()
    return {
        'total_receipts': ExportReceipt.objects.count(),
        'pending_receipts': ExportReceipt.objects.filter(status='PENDING').count(),
        'total_items': ExportReceiptItem.objects.aggregate(total=Sum('quantity'))['total'] or 0,
        'today_transactions': ExportReceipt.objects.filter(created_at__date=today).count(),
    }


PAGE_SIZE = 5  # Số phiếu mỗi trang


def _get_user_display_name(user):
    full_name = ''
    if hasattr(user, 'get_full_name'):
        full_name = (user.get_full_name() or '').strip()
    return full_name or getattr(user, 'username', '') or 'Khong ro'


def _resolve_report_category_label(service, category_id):
    label = 'Tat ca danh muc'
    if category_id:
        selected = next((cat for cat in service.get_categories() if str(cat.id) == category_id), None)
        if selected:
            label = selected.name
    return label


def _format_report_number(value):
    decimal_value = Decimal(str(value))
    formatted = f"{decimal_value:,.10f}".rstrip('0').rstrip('.')
    return formatted or '0'


def _parse_stock_report_filters(request):
    today = timezone.localdate()
    first_day = today.replace(day=1)

    from_date_str = request.GET.get('from_date', first_day.isoformat())
    to_date_str = request.GET.get('to_date', today.isoformat())
    raw_category_id = request.GET.get('category', '').strip()

    category_id = ''
    if raw_category_id:
        try:
            uuid.UUID(raw_category_id)
            category_id = raw_category_id
        except ValueError:
            messages.error(request, 'Danh mục không hợp lệ. Hệ thống đã bỏ bộ lọc danh mục.')

    try:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    except ValueError:
        from_date = first_day
        to_date = today
        messages.error(request, 'Khoảng thời gian không hợp lệ. Hệ thống đã dùng mặc định tháng hiện tại.')

    is_valid_range = True
    if from_date > to_date:
        is_valid_range = False
        messages.error(request, 'Ngày bắt đầu không được lớn hơn ngày kết thúc. Vui lòng chọn lại khoảng thời gian.')

    return from_date, to_date, category_id, is_valid_range


# ═══════════════════════════════════════════════════════════════
# NHẬP KHO
# ═══════════════════════════════════════════════════════════════

class ImportReceiptListView(LoginRequiredMixin, View):
    def get(self, request):
        service = ImportReceiptService()
        user = request.user

        if user.role in ('KE_TOAN', 'ADMIN'):
            receipts = service.get_all()
        else:
            receipts = service.get_by_user(user)

        status_filter = request.GET.get('status', '')
        search_query = request.GET.get('search', '')
        page_number = request.GET.get('page', 1)

        if status_filter:
            receipts = receipts.filter(status=status_filter)

        if search_query:
            receipts = receipts.filter(
                Q(receipt_code__icontains=search_query) |
                Q(note__icontains=search_query)
            )

        paginator = Paginator(receipts, PAGE_SIZE)
        page_obj = paginator.get_page(page_number)

        products_data = _products_json()
        stats = _get_import_receipt_stats()

        user_role = 'ADMIN' if user.is_superuser else user.role

        return render(request, 'warehouse/import_receipt_list.html', {
            'receipts': page_obj,
            'page_obj': page_obj,
            'paginator': paginator,
            'products_json': json.dumps(products_data, ensure_ascii=False),
            'status_filter': status_filter,
            'search_query': search_query,
            'user_role': user_role,
            'stats': stats,
        })

    def post(self, request):
        if request.user.role not in ('KHO', 'ADMIN') and not request.user.is_superuser:
            messages.error(request, 'Bạn không có quyền tạo phiếu nhập kho.')
            return redirect('warehouse:import_list')

        service = ImportReceiptService()
        note = request.POST.get('note', '')
        items_data = _parse_items_from_post(request.POST)

        receipt, error = service.create_receipt(note, items_data, request.user)
        if receipt:
            messages.success(request, f'Đã tạo phiếu nhập {receipt.receipt_code} thành công. Đang chờ kế toán duyệt.')
        else:
            messages.error(request, error)

        return redirect('warehouse:import_list')


class ImportReceiptDetailView(LoginRequiredMixin, View):
    def get(self, request, pk):
        service = ImportReceiptService()
        receipt = service.get_by_id(pk)
        if not receipt:
            messages.error(request, 'Không tìm thấy phiếu nhập kho.')
            return redirect('warehouse:import_list')

        products_data = _products_json()
        return render(request, 'warehouse/import_receipt_detail.html', {
            'receipt': receipt,
            'products_json': json.dumps(products_data, ensure_ascii=False),
            'user_role': 'ADMIN' if request.user.is_superuser else request.user.role,
        })


class ImportReceiptApproveView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if request.user.role not in ('KE_TOAN', 'ADMIN') and not request.user.is_superuser:
            messages.error(request, 'Bạn không có quyền duyệt phiếu.')
            return redirect('warehouse:import_list')

        service = ImportReceiptService()
        success, msg = service.approve_receipt(pk, request.user)
        if success:
            messages.success(request, msg)
        else:
            messages.error(request, msg)
        return redirect('warehouse:import_list')


class ImportReceiptRejectView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if request.user.role not in ('KE_TOAN', 'ADMIN') and not request.user.is_superuser:
            messages.error(request, 'Bạn không có quyền từ chối phiếu.')
            return redirect('warehouse:import_list')

        service = ImportReceiptService()
        rejection_note = request.POST.get('rejection_note', '')
        success, msg = service.reject_receipt(pk, request.user, rejection_note)
        if success:
            messages.warning(request, msg)
        else:
            messages.error(request, msg)
        return redirect('warehouse:import_list')


class ImportReceiptResubmitView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if request.user.role not in ('KHO', 'ADMIN') and not request.user.is_superuser:
            messages.error(request, 'Bạn không có quyền gửi lại phiếu.')
            return redirect('warehouse:import_list')

        service = ImportReceiptService()
        note = request.POST.get('note', '')
        items_data = _parse_items_from_post(request.POST)

        receipt, error = service.resubmit_receipt(pk, note, items_data, request.user)
        if receipt:
            messages.success(request, f'Đã gửi lại phiếu {receipt.receipt_code}. Đang chờ kế toán duyệt.')
        else:
            messages.error(request, error)
        return redirect('warehouse:import_list')


# ═══════════════════════════════════════════════════════════════
# TỒN KHO
# ═══════════════════════════════════════════════════════════════

class StockListView(LoginRequiredMixin, View):
    def get(self, request):
        service = StockService()
        search_query = request.GET.get('search', '')
        page_number = request.GET.get('page', 1)

        stocks_qs = service.get_all_stocks()

        if search_query:
            stocks_qs = stocks_qs.filter(
                Q(product__name__icontains=search_query) |
                Q(product__category__name__icontains=search_query)
            )

        paginator = Paginator(stocks_qs, 12)  # 12 cards mỗi trang
        page_obj = paginator.get_page(page_number)

        return render(request, 'warehouse/stock_list.html', {
            'stocks': page_obj,
            'page_obj': page_obj,
            'paginator': paginator,
            'search_query': search_query,
            'user_role': 'ADMIN' if request.user.is_superuser else request.user.role,
        })


class StockReportView(LoginRequiredMixin, View):
    def get(self, request):
        service = StockReportService()
        from_date, to_date, category_id, is_valid_range = _parse_stock_report_filters(request)

        if is_valid_range:
            rows, totals = service.build_report(
                from_date=from_date,
                to_date=to_date,
                category_id=category_id or None,
            )
        else:
            rows = []
            totals = {
                'opening': 0,
                'import_qty': 0,
                'export_qty': 0,
                'closing': 0,
            }

        return render(request, 'warehouse/stock_report.html', {
            'rows': rows,
            'totals': totals,
            'categories': service.get_categories(),
            'from_date': from_date.isoformat(),
            'to_date': to_date.isoformat(),
            'category_id': category_id,
            'generated_at': timezone.localtime(),
            'generated_by': _get_user_display_name(request.user),
            'user_role': 'ADMIN' if request.user.is_superuser else request.user.role,
        })


class StockReportExportExcelView(LoginRequiredMixin, View):
    def get(self, request):
        service = StockReportService()
        from_date, to_date, category_id, is_valid_range = _parse_stock_report_filters(request)

        if not is_valid_range:
            params = {
                'from_date': from_date.isoformat(),
                'to_date': to_date.isoformat(),
            }
            if category_id:
                params['category'] = category_id
            return redirect(f"{reverse('warehouse:stock_report')}?{urlencode(params)}")

        rows, totals = service.build_report(
            from_date=from_date,
            to_date=to_date,
            category_id=category_id or None,
        )

        # Import tại chỗ để tránh làm lỗi toàn module nếu môi trường chưa cài dependency.
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Bao cao ton kho'

        title = 'BAO CAO TON KHO THEO THOI GIAN'
        category_label = _resolve_report_category_label(service, category_id)

        sheet.merge_cells('A1:H1')
        sheet['A1'] = title
        sheet['A1'].font = Font(size=14, bold=True)
        sheet['A1'].alignment = Alignment(horizontal='center')

        sheet['A2'] = f'Tu ngay: {from_date.strftime("%d/%m/%Y")}'
        sheet['A3'] = f'Den ngay: {to_date.strftime("%d/%m/%Y")}'
        sheet['A4'] = f'Danh muc: {category_label}'
        sheet['A5'] = f'Xuat luc: {timezone.localtime().strftime("%d/%m/%Y %H:%M:%S")}'
        sheet['A6'] = f'Nguoi xuat: {_get_user_display_name(request.user)}'

        headers = [
            'STT',
            'San pham',
            'Danh muc',
            'Don vi',
            'Ton dau ky',
            'Nhap trong ky',
            'Xuat trong ky',
            'Ton cuoi ky',
        ]

        header_row = 7
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(fill_type='solid', fgColor='1F4E78')
            cell.alignment = Alignment(horizontal='center')

        # Không ép hiển thị .00; chỉ hiện phần thập phân khi thật sự có.
        number_format = '#,##0.##'
        start_data_row = header_row + 1
        current_row = start_data_row
        for index, row in enumerate(rows, start=1):
            sheet.cell(row=current_row, column=1, value=index)
            sheet.cell(row=current_row, column=2, value=row['product'].name)
            sheet.cell(row=current_row, column=3, value=row['product'].category.name if row['product'].category else '')
            sheet.cell(row=current_row, column=4, value=row['product'].base_unit)

            opening_cell = sheet.cell(row=current_row, column=5, value=float(row['opening']))
            import_cell = sheet.cell(row=current_row, column=6, value=float(row['import_qty']))
            export_cell = sheet.cell(row=current_row, column=7, value=float(row['export_qty']))
            closing_cell = sheet.cell(row=current_row, column=8, value=float(row['closing']))

            opening_cell.number_format = number_format
            import_cell.number_format = number_format
            export_cell.number_format = number_format
            closing_cell.number_format = number_format

            current_row += 1

        total_row = current_row
        sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
        total_title_cell = sheet.cell(row=total_row, column=1, value='Tong cong')
        total_title_cell.font = Font(bold=True)
        total_title_cell.alignment = Alignment(horizontal='right')

        total_opening = sheet.cell(row=total_row, column=5, value=float(totals['opening']))
        total_import = sheet.cell(row=total_row, column=6, value=float(totals['import_qty']))
        total_export = sheet.cell(row=total_row, column=7, value=float(totals['export_qty']))
        total_closing = sheet.cell(row=total_row, column=8, value=float(totals['closing']))

        for total_cell in (total_opening, total_import, total_export, total_closing):
            total_cell.number_format = number_format
            total_cell.font = Font(bold=True)

        column_widths = {
            'A': 7,
            'B': 32,
            'C': 20,
            'D': 12,
            'E': 15,
            'F': 15,
            'G': 15,
            'H': 15,
        }
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width

        filename = f'bao_cao_ton_kho_{from_date.isoformat()}_{to_date.isoformat()}.xlsx'
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        workbook.save(response)
        return response


class StockReportExportPdfView(LoginRequiredMixin, View):
    def get(self, request):
        service = StockReportService()
        from_date, to_date, category_id, is_valid_range = _parse_stock_report_filters(request)

        if not is_valid_range:
            params = {
                'from_date': from_date.isoformat(),
                'to_date': to_date.isoformat(),
            }
            if category_id:
                params['category'] = category_id
            return redirect(f"{reverse('warehouse:stock_report')}?{urlencode(params)}")

        rows, totals = service.build_report(
            from_date=from_date,
            to_date=to_date,
            category_id=category_id or None,
        )

        # Import tại chỗ để tránh làm lỗi toàn module nếu môi trường chưa cài dependency.
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        category_label = _resolve_report_category_label(service, category_id)
        filename = f'bao_cao_ton_kho_{from_date.isoformat()}_{to_date.isoformat()}.pdf'

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(A4),
            leftMargin=20,
            rightMargin=20,
            topMargin=20,
            bottomMargin=20,
        )

        styles = getSampleStyleSheet()
        regular_font = 'Helvetica'
        bold_font = 'Helvetica-Bold'

        # Ưu tiên font hỗ trợ tiếng Việt nếu có sẵn trong container.
        regular_font_path = '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'
        bold_font_path = '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'
        if os.path.exists(regular_font_path) and os.path.exists(bold_font_path):
            pdfmetrics.registerFont(TTFont('DejaVuSans', regular_font_path))
            pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', bold_font_path))
            regular_font = 'DejaVuSans'
            bold_font = 'DejaVuSans-Bold'

        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            alignment=TA_CENTER,
            fontName=bold_font,
            fontSize=16,
            spaceAfter=10,
        )
        body_style = ParagraphStyle(
            'ReportBody',
            parent=styles['Normal'],
            fontName=regular_font,
            fontSize=10,
        )

        story = [
            Paragraph('BAO CAO TON KHO THEO THOI GIAN', title_style),
            Paragraph(f'Tu ngay: {from_date.strftime("%d/%m/%Y")}', body_style),
            Paragraph(f'Den ngay: {to_date.strftime("%d/%m/%Y")}', body_style),
            Paragraph(f'Danh muc: {category_label}', body_style),
            Paragraph(f'Xuat luc: {timezone.localtime().strftime("%d/%m/%Y %H:%M:%S")}', body_style),
            Paragraph(f'Nguoi xuat: {_get_user_display_name(request.user)}', body_style),
            Spacer(1, 10),
        ]

        def fmt(value):
            return _format_report_number(value)

        table_data = [[
            'STT',
            'San pham',
            'Danh muc',
            'Don vi',
            'Ton dau ky',
            'Nhap trong ky',
            'Xuat trong ky',
            'Ton cuoi ky',
        ]]

        for index, row in enumerate(rows, start=1):
            table_data.append([
                str(index),
                str(row['product'].name),
                str(row['product'].category.name if row['product'].category else ''),
                str(row['product'].base_unit),
                fmt(row['opening']),
                fmt(row['import_qty']),
                fmt(row['export_qty']),
                fmt(row['closing']),
            ])

        table_data.append([
            '',
            'TONG CONG',
            '',
            '',
            fmt(totals['opening']),
            fmt(totals['import_qty']),
            fmt(totals['export_qty']),
            fmt(totals['closing']),
        ])

        table = Table(
            table_data,
            repeatRows=1,
            colWidths=[30, 180, 120, 70, 85, 85, 85, 85],
        )
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), bold_font),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
            ('FONTNAME', (0, 1), (-1, -2), regular_font),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F3F6FA')),
            ('FONTNAME', (0, -1), (-1, -1), bold_font),
        ])
        table.setStyle(table_style)

        story.append(table)
        doc.build(story)
        return response


# ═══════════════════════════════════════════════════════════════
# XUẤT KHO — Tất cả role đều có thể duyệt
# ═══════════════════════════════════════════════════════════════

EXPORT_APPROVE_ROLES = ('KHO', 'KE_TOAN', 'ADMIN', 'SALE')


class ExportReceiptListView(LoginRequiredMixin, View):
    def get(self, request):
        service = ExportReceiptService()
        user = request.user

        receipts = service.get_all()

        status_filter = request.GET.get('status', '')
        search_query = request.GET.get('search', '')
        page_number = request.GET.get('page', 1)

        if status_filter:
            receipts = receipts.filter(status=status_filter)

        if search_query:
            receipts = receipts.filter(
                Q(receipt_code__icontains=search_query) |
                Q(note__icontains=search_query)
            )

        paginator = Paginator(receipts, PAGE_SIZE)
        page_obj = paginator.get_page(page_number)

        products_data = _products_json()
        stats = _get_export_receipt_stats()

        return render(request, 'warehouse/export_receipt_list.html', {
            'receipts': page_obj,
            'page_obj': page_obj,
            'paginator': paginator,
            'products_json': json.dumps(products_data, ensure_ascii=False),
            'status_filter': status_filter,
            'search_query': search_query,
            'user_role': 'ADMIN' if user.is_superuser else user.role,
            'stats': stats,
        })

    def post(self, request):
        if request.user.role not in ('KHO', 'ADMIN') and not request.user.is_superuser:
            messages.error(request, 'Bạn không có quyền tạo phiếu xuất kho.')
            return redirect('warehouse:export_list')

        service = ExportReceiptService()
        note = request.POST.get('note', '')
        items_data = _parse_items_from_post(request.POST)

        receipt, error = service.create_receipt(note, items_data, request.user)
        if receipt:
            messages.success(request, f'Đã tạo phiếu xuất {receipt.receipt_code} thành công. Chờ duyệt.')
        else:
            messages.error(request, error)

        return redirect('warehouse:export_list')


class ExportReceiptDetailView(LoginRequiredMixin, View):
    def get(self, request, pk):
        service = ExportReceiptService()
        receipt = service.get_by_id(pk)
        if not receipt:
            messages.error(request, 'Không tìm thấy phiếu xuất kho.')
            return redirect('warehouse:export_list')

        products_data = _products_json()
        return render(request, 'warehouse/export_receipt_detail.html', {
            'receipt': receipt,
            'products_json': json.dumps(products_data, ensure_ascii=False),
            'user_role': 'ADMIN' if request.user.is_superuser else request.user.role,
        })


class ExportReceiptApproveView(LoginRequiredMixin, View):
    def post(self, request, pk):
        service = ExportReceiptService()
        success, msg = service.approve_receipt(pk, request.user)
        if success:
            messages.success(request, msg)
        else:
            messages.error(request, msg)
        return redirect('warehouse:export_list')


class ExportReceiptRejectView(LoginRequiredMixin, View):
    def post(self, request, pk):
        service = ExportReceiptService()
        rejection_note = request.POST.get('rejection_note', '')
        success, msg = service.reject_receipt(pk, request.user, rejection_note)
        if success:
            messages.warning(request, msg)
        else:
            messages.error(request, msg)
        return redirect('warehouse:export_list')


class ExportReceiptResubmitView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if request.user.role not in ('KHO', 'ADMIN') and not request.user.is_superuser:
            messages.error(request, 'Bạn không có quyền gửi lại phiếu.')
            return redirect('warehouse:export_list')

        service = ExportReceiptService()
        note = request.POST.get('note', '')
        items_data = _parse_items_from_post(request.POST)

        receipt, error = service.resubmit_receipt(pk, note, items_data, request.user)
        if receipt:
            messages.success(request, f'Đã gửi lại phiếu {receipt.receipt_code}. Chờ duyệt.')
        else:
            messages.error(request, error)
        return redirect('warehouse:export_list')